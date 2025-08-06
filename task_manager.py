"""
This script manages recurring tasks in the `ProductivitySystem.xlsx` workbook.
It reads tasks from the **Tasks** sheet, identifies completed tasks, creates
new tasks based on their recurrence pattern, and archives completed tasks
in the **Logs** sheet.

Usage:
    python task_manager.py --file /path/to/ProductivitySystem.xlsx

The **Tasks** sheet must have columns in the following order:
1. Task ID (integer)
2. Task Name (string)
3. Due Date (Excel date)
4. Status ("Pending" or "Done")
5. Recurrence ("Daily", "Weekly", "Monthly", or empty)
6. Notes (string)

Completed tasks with a recurrence value will have a new task created with
an updated due date. For Monthly tasks, the next due date is approximated
by adding 30 days.
"""

import argparse
from datetime import datetime, timedelta
from openpyxl import load_workbook


def process_tasks(workbook_path: str) -> None:
    """Process recurring tasks in the given workbook.

    Args:
        workbook_path: Path to the Excel workbook (XLSX file).
    """
    wb = load_workbook(workbook_path)
    tasks_ws = wb["Tasks"]
    logs_ws = wb["Logs"]

    # Determine the next task ID
    existing_ids = [cell.value for cell in tasks_ws["A"] if isinstance(cell.value, int)]
    next_id = max(existing_ids) + 1 if existing_ids else 1

    today = datetime.today().date()
    rows_to_delete = []

    for row_index, row in enumerate(tasks_ws.iter_rows(min_row=2, values_only=False), start=2):
        task_id_cell, name_cell, due_cell, status_cell, recurrence_cell, notes_cell = row
        status = status_cell.value
        recurrence = recurrence_cell.value
        due_date = due_cell.value

        # Skip rows without a valid due date
        if due_date is None or not hasattr(due_date, "date"):
            continue
        due_date = due_date.date()

        # Only process tasks marked as Done
        if status and str(status).lower() == "done":
            # Determine next due date based on recurrence
            next_due = None
            if recurrence:
                rec_lower = str(recurrence).lower()
                if rec_lower == "daily":
                    next_due = due_date + timedelta(days=1)
                elif rec_lower == "weekly":
                    next_due = due_date + timedelta(weeks=1)
                elif rec_lower == "monthly":
                    # Approximate one month as 30 days
                    next_due = due_date + timedelta(days=30)

            # Append completed task to Logs
            logs_ws.append([
                task_id_cell.value,
                name_cell.value,
                today,
                due_date,
                notes_cell.value,
            ])

            # Create a new task if recurrence applies
            if next_due:
                tasks_ws.append([
                    next_id,
                    name_cell.value,
                    next_due,
                    "Pending",
                    recurrence_cell.value,
                    notes_cell.value,
                ])
                next_id += 1

            rows_to_delete.append(row_index)

    # Delete processed rows from bottom to top
    for idx in sorted(rows_to_delete, reverse=True):
        tasks_ws.delete_rows(idx)

    wb.save(workbook_path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Process recurring tasks in the productivity workbook.")
    parser.add_argument("--file", required=True, help="Path to the Excel workbook.")
    args = parser.parse_args()
    process_tasks(args.file)


if __name__ == "__main__":
    main()
